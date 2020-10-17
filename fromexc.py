from openpyxl import Workbook, load_workbook

try:
    wb = load_workbook(filename='4217.xlsx')
except FileNotFoundError:
    print('>>> problem with file')

sheet_val = wb['Sheet1']
sheet_ranges = wb.sheetnames
wb2 = load_workbook(filename='test_intex.xlsx')
sheet_val2 = wb2['Лист1']


def get_row(target_row, sheet_name, position):
    """ возвращает список данных в столбце
    target_row: str
    sheet_name: workbook['sheet_name']
    position: start line
    """
    row_data = []
    i = position
    while i:
        if not sheet_name[f'{target_row}{i}'].value:
            break
        data = sheet_name[f'{target_row}{i}'].value
        i += 1
        row_data.append(data)
    return row_data


# print(get_row('c', sheet_val, 14))
# print(get_row('d', sheet_val2, 1))
data_in = get_row('c', sheet_val, 14)
data_out = get_row('d', sheet_val2, 1)
print(f'data_in кол-во: {len(data_in)}, data_out кол-во: {len(data_out)}')


for item in data_in:
    for index in data_out:
        if item.find(str(index)) == 1:
            data_out.remove(index)

print(f'>>> нет в списке:\n {data_out}')
